from openpyxl import load_workbook

wb = load_workbook(filename='store_data.xlsx', data_only=True)

# Load worksheets
customer_sheet = wb['Customers']
products_sheet = wb['Products']
invoice_sheet = wb['Invoices']
invoiceItems_sheet = wb['Invoice_details']

# for i in range(1, invoice_sheet.max_row):
#     header = invoice_sheet.cell(row=i, column=1).value
#     # invoice_id = invoice_sheet.cell(row=i, column=1).value
#     # customer_id = invoice_sheet.cell(row=i, column=2).value
#     # order_date = invoice_sheet.cell(row=i, column=3).value
#     # invoice_amt = invoice_sheet.cell(row=i, column=4).value
#     # shipping_add = invoice_sheet.cell(row=i, column=5).value
#     # print("{:.2f}".format(invoice_amt))
#     print(header)

for i in range(1, invoice_sheet.max_row):
    for j in range(1, invoice_sheet.max_column):
        cell_obj = invoice_sheet.cell(row=i, column=j)
        print(cell_obj.value, end=' | ')
    print('\n')


# Get invoice id to be generated as input from user
# target_invoice = input('Enter Invoice Id: ')
# try:
#     target_invoice = int(target_invoice)
# except:
#     print("Invalid input, terminating...")
# if target_invoice < 101:
#     print("Invalid Invoice ID, terminating...")


# for row in invoice_sheet.iter_rows(invoice_sheet.min_row, invoice_sheet.max_row):
#     for cell in row:
#         print(cell.value)

# for row in invoice_sheet.iter_rows(min_row=1, max_col=1, values_only=True):
#     for value in row:
#         print(value)

# Find row of the invoice id provided
# Collect all data in the row
# Access the cells required for display on the invoice file

# print(c1)
# print(target_invoice)
