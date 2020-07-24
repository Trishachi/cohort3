from openpyxl import load_workbook

wb = load_workbook(filename='store_data.xlsx', data_only=True)

# Load worksheets
customer_sheet = wb['Customers']
products_sheet = wb['Products']
invoice_sheet = wb['Invoices']
invoiceItems_sheet = wb['Invoice_details']

# Function for creating dictionaries


def makeDict(sheet):
    header = sheet[1]
    newDict = {}
    for row in sheet.rows:
        rowDict = {}
        index = 0
        for cell in row:
            rowDict[header[index].value] = cell.value
            index = index+1
        newDict[row[0].value] = rowDict
    # newDict.pop('Customer_id')
    return newDict


def createXl_invoice(invoiceNumber, invoiceDate, customerID,
                     shippingAddress, invoiceTotal, customerFullname, customerEmail, customerAdd, customerPhone,
                     item1_name, item1_price, item1_qty, item2_name, item2_price, item2_qty,
                     item3_name, item3_price, item3_qty):
    invoice = load_workbook(filename='invoice_template.xlsx', data_only=True)
    current_invoice = invoice.active
    current_invoice['H4'] = invoiceDate
    current_invoice['H6'] = invoiceNumber
    current_invoice['H8'] = customerID
    current_invoice['D14'] = shippingAddress
    current_invoice['F14'] = customerAdd
    current_invoice['H35'] = invoiceTotal
    current_invoice['D12'] = customerFullname
    current_invoice['F12'] = customerFullname
    current_invoice['D15'] = customerPhone
    current_invoice['F15'] = customerPhone
    current_invoice['D16'] = customerEmail

    current_invoice['D20'] = item1_name
    current_invoice['F20'] = item1_price
    current_invoice['G20'] = item1_qty
    current_invoice['D21'] = item2_name
    current_invoice['F21'] = item2_price
    current_invoice['G21'] = item2_qty
    current_invoice['D22'] = item3_name
    current_invoice['F22'] = item3_price
    current_invoice['G22'] = item3_qty

    current_invoice['H20'] = item1_price * item1_qty
    current_invoice['H21'] = item2_price * item2_qty
    current_invoice['H22'] = item3_price * item3_qty

    current_invoice['H29'] = current_invoice['H20'].value + \
        current_invoice['H21'].value + current_invoice['H22'].value

    invoice.save('invoice_template.xlsx')


# Create dictionaries from sheets
customerDict = makeDict(customer_sheet)
invoicesDict = makeDict(invoice_sheet)
productsDict = makeDict(products_sheet)
invoiceItemsDict = makeDict(invoiceItems_sheet)

# Get invoice id to be generated as input from user
target_invoice = input('Enter Invoice Id: ')
try:
    target_invoice = int(target_invoice)
except:
    print("Invalid input, terminating...")
if target_invoice < 101:
    print("Invalid Invoice ID, terminating...")

# Find and get data of row of the invoice id provided
targetInvoice_row = invoicesDict[target_invoice]

# Access the cells required for display on the invoice file
invoiceNumber = targetInvoice_row.get("Invoice_id")
invoiceDate = targetInvoice_row.get("Order_date")
customerID = targetInvoice_row.get("Customer_id")
shippingAddress = targetInvoice_row.get("Shipping_add")
invoiceTotal = targetInvoice_row.get("Invoice_amt")

targetCustomer_row = customerDict[customerID]

customerFirstname = targetCustomer_row.get("First_name")
customerLastname = targetCustomer_row.get("Last_name")
customerFullname = customerFirstname + " " + customerLastname
customerEmail = targetCustomer_row.get("Email")
customerAdd = targetCustomer_row.get("Address")
customerPhone = targetCustomer_row.get("Phone")

itemsDict = {}
rowDict = {}
counter = 1

for row in invoiceItems_sheet.values:
    if row[0] == target_invoice:
        rowDict = {'Item': counter,
                   'Name': row[2], 'Price': row[3], 'Qty': row[4]}
        itemsDict[counter] = rowDict
        counter = counter + 1

# print(itemsDict)

item1_row = itemsDict[1]
item2_row = itemsDict[2]
item3_row = itemsDict[3]

item1_name = item1_row.get("Name")
item1_price = item1_row.get("Price")
item1_qty = item1_row.get("Qty")
item2_name = item2_row.get("Name")
item2_price = item2_row.get("Price")
item2_qty = item2_row.get("Qty")
item3_name = item3_row.get("Name")
item3_price = item3_row.get("Price")
item3_qty = item3_row.get("Qty")


createXl_invoice(invoiceNumber, invoiceDate, customerID,
                 shippingAddress, invoiceTotal, customerFullname, customerEmail, customerAdd, customerPhone,
                 item1_name, item1_price, item1_qty, item2_name, item2_price, item2_qty,
                 item3_name, item3_price, item3_qty)
